import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import os

# Define the path for the Excel file
excel_file = 'inventory.xlsx'

# Initialize the Excel file if it doesn't exist
def init_excel():
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Product ID", "Product Name", "Product Description", "Price", "Quantity"])
        wb.save(excel_file)

# Function to add a product
def add_product():
    product_id = entry_id.get()
    name = entry_name.get()
    description = entry_description.get()
    price = entry_price.get()
    quantity = entry_quantity.get()
    
    if product_id and name and description and price and quantity:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append([product_id, name, description, price, quantity])
        wb.save(excel_file)
        messagebox.showinfo("Success", "Product added successfully!")
        clear_entries()
    else:
        messagebox.showwarning("Input Error", "Please fill all fields")

# Function to search for a product by name
def search_product():
    name = entry_name.get()
    if not name:
        messagebox.showwarning("Input Error", "Please enter the product name to search")
        return
    
    wb = load_workbook(excel_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == name:
            entry_id.delete(0, tk.END)
            entry_id.insert(0, row[0])
            entry_description.delete(0, tk.END)
            entry_description.insert(0, row[2])
            entry_price.delete(0, tk.END)
            entry_price.insert(0, row[3])
            entry_quantity.delete(0, tk.END)
            entry_quantity.insert(0, row[4])
            return
    messagebox.showwarning("Not Found", "Product not found")

# Function to update a product
def update_product():
    product_id = entry_id.get()
    name = entry_name.get()
    description = entry_description.get()
    price = entry_price.get()
    quantity = entry_quantity.get()
    
    if product_id and name and description and price and quantity:
        wb = load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product_id:
                row[1].value = name
                row[2].value = description
                row[3].value = price
                row[4].value = quantity
                wb.save(excel_file)
                messagebox.showinfo("Success", "Product updated successfully!")
                clear_entries()
                return
        messagebox.showwarning("Not Found", "Product not found")
    else:
        messagebox.showwarning("Input Error", "Please fill all fields")

# Function to clear entries
def clear_entries():
    entry_id.delete(0, tk.END)
    entry_name.delete(0, tk.END)
    entry_description.delete(0, tk.END)
    entry_price.delete(0, tk.END)
    entry_quantity.delete(0, tk.END)

# Function to view all inventory
def view_inventory():
    wb = load_workbook(excel_file)
    ws = wb.active
    inventory_window = tk.Toplevel(root)
    inventory_window.title("Current Inventory")
    tree = ttk.Treeview(inventory_window, columns=("Product ID", "Product Name", "Product Description", "Price", "Quantity"), show='headings')
    tree.heading("Product ID", text="Product ID")
    tree.heading("Product Name", text="Product Name")
    tree.heading("Product Description", text="Product Description")
    tree.heading("Price", text="Price")
    tree.heading("Quantity", text="Quantity")

    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert('', tk.END, values=row)
    
    tree.pack(fill=tk.BOTH, expand=True)

# Function to sell a product
def sell_product():
    sell_window = tk.Toplevel(root)
    sell_window.title("Sell Product")

    tk.Label(sell_window, text="Product Name").grid(row=0, column=0, padx=10, pady=5)
    sell_entry_name = tk.Entry(sell_window)
    sell_entry_name.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(sell_window, text="Quantity to Sell").grid(row=1, column=0, padx=10, pady=5)
    sell_entry_quantity = tk.Entry(sell_window)
    sell_entry_quantity.grid(row=1, column=1, padx=10, pady=5)

    def confirm_sell():
        name = sell_entry_name.get()
        quantity_to_sell = sell_entry_quantity.get()

        if not name or not quantity_to_sell:
            messagebox.showwarning("Input Error", "Please enter the product name and quantity to sell")
            return
        
        quantity_to_sell = int(quantity_to_sell)
        wb = load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[1].value == name:
                current_quantity = int(row[4].value)
                if current_quantity >= quantity_to_sell:
                    row[4].value = current_quantity - quantity_to_sell
                    wb.save(excel_file)
                    messagebox.showinfo("Success", "Product sold successfully!")
                    clear_entries()
                    sell_window.destroy()
                    return
                else:
                    messagebox.showwarning("Error", "Not enough quantity in stock")
                    return
        messagebox.showwarning("Not Found", "Product not found")
    
    tk.Button(sell_window, text="CONFIRM SELL", command=confirm_sell, bg='orange').grid(row=2, column=1, padx=10, pady=5)

# Function to delete a product
def delete_product():
    delete_window = tk.Toplevel(root)
    delete_window.title("Delete Product")

    tk.Label(delete_window, text="Product ID").grid(row=0, column=0, padx=10, pady=5)
    delete_entry_id = tk.Entry(delete_window)
    delete_entry_id.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(delete_window, text="Product Name").grid(row=1, column=0, padx=10, pady=5)
    delete_entry_name = tk.Entry(delete_window)
    delete_entry_name.grid(row=1, column=1, padx=10, pady=5)

    def confirm_delete():
        product_id = delete_entry_id.get()
        name = delete_entry_name.get()

        if not product_id and not name:
            messagebox.showwarning("Input Error", "Please enter the product ID or product name to delete")
            return
        
        wb = load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product_id or row[1].value == name:
                ws.delete_rows(row[0].row, 1)
                wb.save(excel_file)
                messagebox.showinfo("Success", "Product deleted successfully!")
                delete_window.destroy()
                return
        messagebox.showwarning("Not Found", "Product not found")

    tk.Button(delete_window, text="CONFIRM DELETE", command=confirm_delete, bg='red').grid(row=2, column=1, padx=10, pady=5)

# Initialize the Excel file
init_excel()

# Create the main window
root = tk.Tk()
root.title("HAQ BAHOO WHOLESALLER")

# Create and place labels and entry fields
tk.Label(root, text="Product ID").grid(row=0, column=0, padx=10, pady=5)
entry_id = tk.Entry(root)
entry_id.grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Product Name").grid(row=1, column=0, padx=10, pady=5)
entry_name = tk.Entry(root)
entry_name.grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="Product Description").grid(row=2, column=0, padx=10, pady=5)
entry_description = tk.Entry(root)
entry_description.grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="Price").grid(row=3, column=0, padx=10, pady=5)
entry_price = tk.Entry(root)
entry_price.grid(row=3, column=1, padx=10, pady=5)

tk.Label(root, text="Quantity").grid(row=4, column=0, padx=10, pady=5)
entry_quantity = tk.Entry(root)
entry_quantity.grid(row=4, column=1, padx=10, pady=5)

# Create and place buttons
tk.Button(root, text="ADD", command=add_product, bg='cyan').grid(row=5, column=0, padx=10, pady=5)
tk.Button(root, text="SEARCH", command=search_product, bg='yellow').grid(row=5, column=1, padx=10, pady=5)
tk.Button(root, text="UPDATE", command=update_product, bg='blue').grid(row=5, column=2, padx=10, pady=5)
tk.Button(root, text="CLEAR", command=clear_entries, bg='red').grid(row=5, column=3, padx=10, pady=5)
tk.Button(root, text="VIEW INVENTORY", command=view_inventory, bg='green').grid(row=6, column=0, padx=10, pady=5)
tk.Button(root, text="SELL", command=sell_product, bg='orange').grid(row=6, column=1, padx=10, pady=5)
tk.Button(root, text="DELETE", command=delete_product, bg='purple').grid(row=6, column=2, padx=10, pady=5)

# Run the application
root.mainloop()
